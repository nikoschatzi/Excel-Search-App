import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from fuzzywuzzy import fuzz
import tkinter
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from tkinter.filedialog import askopenfile
from tkinter import filedialog
from tkinter import font as tkFont  # for convenience
from openpyxl.styles import PatternFill
import tkinter as tk
import tkinter.messagebox
import warnings
from tkinter import ttk
import customtkinter
import emoji
from threading import *
warnings.simplefilter("ignore")

global happy_created, runs, begin_gif, ignore_line, disable_gif, stop_line, success
happy_created = 0
runs = 0
begin_gif = 0
problem = 0
ignore_line = 0
disable_gif = 0
stop_line = 100000
success = 100000

#1 --> GUI works, 0 --> GUI doesn't work - my default folders
GUI = 1

def run_code(happy_created=None, label2=None):

    problem = 0
    if GUI == 1:
        request_location = textBox1.get("1.0","end-1c")
        database_location = textBox2.get("1.0","end-1c")
        destination = textBox3.get("1.0","end-1c")
        match_score = textBox4.get("1.0","end-1c")
        stop_line = textBox9.get("1.0","end-1c")
        check_score = match_score.isnumeric()
        check_stop_line = stop_line.isnumeric()
        last_req_chars = request_location[-5:]
        last_data_chars = database_location[-5:]
        choice1 = drop1.get()
        choice2 = drop2.get()
        choice3 = drop3.get()
        choice4 = drop4.get()
        string1 = textBox5.get("1.0","end-1c")
        string2 = textBox6.get("1.0","end-1c")
        string3 = textBox7.get("1.0","end-1c")
        string4 = textBox8.get("1.0","end-1c")

        if request_location == "":
            messagebox.showerror('Error', 'Request file not found!')
            problem = 1

        if database_location == "":
            messagebox.showerror('Error', 'Database file not found!')
            problem = 1

        if last_req_chars != ".xlsx":
            messagebox.showerror('Error', 'Please enter .xlsx request file!')
            problem = 1

        if last_data_chars != ".xlsx":
            messagebox.showerror('Error', 'Please enter .xlsx database file!')
            problem = 1

        if destination == "":
            messagebox.showerror('Error', 'Destination folder not found!')
            problem = 1

        if check_stop_line == False and ignore_line==1:
            messagebox.showerror('Error', 'Stop line is not integer!')
            problem = 1

        if check_score == False:
            messagebox.showerror('Error', 'Match score is not integer!')
            problem = 1

        if match_score == "":
            messagebox.showerror('Error', 'Match score not found!')
            problem = 1

        if choice1 == choice2 or choice1 == choice3 or choice1 == choice4 or choice2 == choice3 or choice2 == choice4 or choice3 == choice4:
            messagebox.showerror('Error', 'Same results columns selected')
            problem = 1

        if string1 == "" and string2 == "" and string3 == "":
            messagebox.showerror('Error', 'Please insert at least one request search string')
            problem = 1

        if string4 == "":
            messagebox.showerror('Error', 'Please insert database search string')
            problem = 1

    if problem == 0:

        t1=Thread(target=run_main)
        t1.start()
        if happy_created == 1:
            label2.destroy()
            happy_created = 0
        global begin_gif
        begin_gif = 1


def run_main():

    global success
    success = 100000
    label_9.grid_remove()
    label_10.grid_remove()
    label_20.grid_remove()
    label_8.grid(row=11, column=3)

    frameCnt = 16
    frames = [PhotoImage(file='work.gif',format = 'gif -index %i' %(i)) for i in range(frameCnt)]
    frameCnt2 = 16
    frames2 = [PhotoImage(file='happy.gif',format = 'gif -index %i' %(i)) for i in range(frameCnt2)]
    frameCnt3 = 16
    frames3 = [PhotoImage(file='fail.gif',format = 'gif -index %i' %(i)) for i in range(frameCnt3)]
    global flag_gif
    flag_gif=1
    global times
    times=0

    def update(ind=None):

        global disable_gif
        if flag_gif==1 and disable_gif==0:
            frame = frames[ind]
            ind += 1
            if ind == frameCnt:
                ind = 0
            #if (times+runs) == 0:
            label.configure(image=frame)
            root.after(100, update, ind)
            return
        if flag_gif==0 and success==1 and disable_gif==0:
            label.destroy()
            label3.destroy()
            happy_created = 1
            if ind >= frameCnt2:
                ind = 0
            frame2 = frames2[ind]
            ind += 1
            #if (times+runs) == 2:
            label2.configure(image=frame2)
            root.after(200, update, ind)
            return
        if (flag_gif==0 and success==0) and disable_gif==0:
            label.destroy()
            label2.destroy()
            happy_created = 1
            if ind >= frameCnt3:
                ind = 0
            frame3 = frames3[ind]
            ind += 1
            #if (times+runs) == 2:
            label3.configure(image=frame3)
            root.after(200, update, ind)
            return
        if disable_gif == 1:
            photo2 = PhotoImage(file="disabled.png")
            labelcern2 = Label(root, image = photo2)
            labelcern2.image = photo2
            labelcern2.grid(row=12, column=2, rowspan=2, columnspan=3)
        else:
            label.destroy()
            label2.destroy()
            label3.destroy()
            return

    label3 = Label(root)
    label3.grid(row=12, column=2, rowspan=2, columnspan=3)
    label2 = Label(root)
    label2.grid(row=12, column=2, rowspan=2, columnspan=3)
    label = Label(root)
    label.grid(row=12, column=2, rowspan=2, columnspan=3)

    if disable_gif==0:
        root.after(0, update, 0)

    #Read data, create new excel
    if GUI ==1:
        stop_line = 100000
        request_location = textBox1.get("1.0","end-1c")
        #print(request_location)
        database_location = textBox2.get("1.0","end-1c")
        destination = textBox3.get("1.0","end-1c")
        match_score = textBox4.get("1.0","end-1c")
        string1 = textBox5.get("1.0","end-1c")
        string2 = textBox6.get("1.0","end-1c")
        string3 = textBox7.get("1.0","end-1c")
        string4 = textBox8.get("1.0","end-1c")

        choice1 = drop1.get()
        choice2 = drop2.get()
        choice3 = drop3.get()
        choice4 = drop4.get()
        string5 = textBox20.get("1.0","end-1c")
        string6 = textBox21.get("1.0","end-1c")
        string7 = textBox22.get("1.0","end-1c")
        string8 = textBox23.get("1.0","end-1c")

        if choice1=="5":
            column_5 = string5
            column_5_select = 0
        if choice2=="5":
            column_5 = string6
            column_5_select = 0
        if choice3=="5":
            column_5 = string7
            column_5_select = 1
        if choice4=="5":
            column_5 = string8
            column_5_select = 1

        if choice1=="7":
            column_7 = string5
            column_7_select = 0
        if choice2=="7":
            column_7 = string6
            column_7_select = 0
        if choice3=="7":
            column_7 = string7
            column_7_select = 1
        if choice4=="7":
            column_7 = string8
            column_7_select = 1

        if choice1=="8":
            column_8 = string5
            column_8_select = 0
        if choice2=="8":
            column_8 = string6
            column_8_select = 0
        if choice3=="8":
            column_8 = string7
            column_8_select = 1
        if choice4=="8":
            column_8 = string8
            column_8_select = 1

        if choice1=="9":
            column_9 = string5
            column_9_select = 0
        if choice2=="9":
            column_9 = string6
            column_9_select = 0
        if choice3=="9":
            column_9 = string7
            column_9_select = 1
        if choice4=="9":
            column_9 = string8
            column_9_select = 1

        if ignore_line == 1:
            stop_line = textBox9.get("1.0","end-1c")

        database_data = load_workbook(database_location)
        request_data = load_workbook(request_location)

    if GUI ==0:
        stop_line = 100000
        match_score = textBox4.get("1.0","end-1c")
        string1 = textBox5.get("1.0","end-1c")
        string2 = textBox6.get("1.0","end-1c")
        string3 = textBox7.get("1.0","end-1c")
        string4 = textBox8.get("1.0","end-1c")
        database_data = load_workbook(r'C:\Users\nchatzip\cernbox\WINDOWS\Desktop\Store_36.xlsx')
        request_data = load_workbook(r'C:\Users\nchatzip\cernbox\WINDOWS\Desktop\altium.xlsx')

        if string1 == "" and string2 == "" and string3 == "":
            messagebox.showinfo('Warning', 'Please insert at least one request search string')

        if string4 == "":
            messagebox.showinfo('Warning', 'Please insert database search string')

    results_data = Workbook()
    not_found_data = Workbook()

    #database_sheet = database_data['Sheet1']
    database_sheet = database_data.worksheets[0]
    request_sheet = request_data.worksheets[0]
    results_sheet = results_data.active

    not_found_sheet = not_found_data.active

    results_sheet.append(['Line', r" Requested {}".format(string1),  r" Requested {}".format(string2), r" Requested {}".format(string3), r"{}".format(column_5), r" Database {}".format(string4), r"{}".format(column_7), r"{}".format(column_8), r"{}".format(column_9), 'Score'])

    not_found_sheet.append(['Line', r"{}".format(string1),  r"{}".format(string2), r"{}".format(string3), r"{}".format(string5), r"{}".format(string6)])

    #Fonts, width
    results_sheet.column_dimensions['A'].width = 5
    results_sheet.column_dimensions['B'].width = 45
    results_sheet.column_dimensions['C'].width = 25
    results_sheet.column_dimensions['D'].width = 25
    results_sheet.column_dimensions['E'].width = 15
    results_sheet.column_dimensions['F'].width = 62
    results_sheet.column_dimensions['G'].width = 10
    results_sheet.column_dimensions['H'].width = 10
    results_sheet.column_dimensions['I'].width = 10
    results_sheet.column_dimensions['J'].width = 6

    not_found_sheet.column_dimensions['A'].width = 7
    not_found_sheet.column_dimensions['B'].width = 50
    not_found_sheet.column_dimensions['C'].width = 25
    not_found_sheet.column_dimensions['D'].width = 25
    not_found_sheet.column_dimensions['E'].width = 20
    not_found_sheet.column_dimensions['F'].width = 12

    results_sheet.row_dimensions[1].height = 25

    not_found_sheet.row_dimensions[1].height = 25

    results_sheet['A1'].font = Font(bold=True)
    results_sheet['B1'].font = Font(bold=True)
    results_sheet['C1'].font = Font(bold=True)
    results_sheet['D1'].font = Font(bold=True)
    results_sheet['E1'].font = Font(bold=True)
    results_sheet['F1'].font = Font(bold=True)
    results_sheet['G1'].font = Font(bold=True)
    results_sheet['H1'].font = Font(bold=True)
    results_sheet['I1'].font = Font(bold=True)
    results_sheet['J1'].font = Font(bold=True)

    not_found_sheet['A1'].font = Font(bold=True)
    not_found_sheet['B1'].font = Font(bold=True)
    not_found_sheet['C1'].font = Font(bold=True)
    not_found_sheet['D1'].font = Font(bold=True)
    not_found_sheet['E1'].font = Font(bold=True)
    not_found_sheet['F1'].font = Font(bold=True)

    #find QIP, description and quantity columns

    not_col_1 = 100000
    not_col_2 = 100000

    for a in request_sheet.iter_rows():
        for b in request_sheet.iter_cols():
            col_no = b[0].column -1
            compare = a[col_no].value

            if compare == string5 and string5 != "":
                not_col_1 = col_no
                break

    for a in request_sheet.iter_rows():
        for b in request_sheet.iter_cols():
            col_no = b[0].column -1
            compare = a[col_no].value

            if compare == string6 and string6 != "":
                not_col_2 = col_no
                break

    quantity_col_request = 100000
    if column_8_select == 0:
        for a in request_sheet.iter_rows():
            for b in request_sheet.iter_cols():
                col_no = b[0].column -1
                compare = a[col_no].value

                if compare == column_8 and column_8 != "":
                    quantity_col_request = col_no
                    break

    if column_8_select == 1:
        for a in database_sheet.iter_rows():
            for b in database_sheet.iter_cols():
                col_no = b[0].column -1
                compare = a[col_no].value

                if compare == column_8 and column_8 != "":
                    quantity_col_request = col_no
                    break

    description_col_request = 100000
    description_row_request = 100000
    details_col_request = 100000
    partno_col_request = 100000
    case_col_request = 100000

    if column_5_select == 0:
        for a in request_sheet.iter_rows():
            for b in request_sheet.iter_cols():
                col_no = b[0].column -1
                compare = a[col_no].value

                if compare == column_5 and column_5 != "":
                    case_col_request = col_no
                    break

    if column_5_select == 1:
        for a in database_sheet.iter_rows():
            for b in database_sheet.iter_cols():
                col_no = b[0].column -1
                compare = a[col_no].value

                if compare == column_5 and column_5 != "":
                    case_col_request = col_no
                    break

    for a in request_sheet.iter_rows():
        row_no = a[0].row
        for b in request_sheet.iter_cols():
            col_no = b[0].column -1
            compare = a[col_no].value

            if string2 == compare and string2 != "" and string2 != " ":
                description_col_request = col_no
                description_row_request = row_no

            if string1 == compare and string1 != "" and string1 != " ":
                details_col_request = col_no
                description_row_request = row_no

            if string3 == compare and string3 != "" and string3 != " ":
                partno_col_request = col_no
                description_row_request = row_no

    if (string1 != "") and details_col_request == 100000:
        messagebox.showinfo('Info',  r"Requested column '{}' was not found. Check spaces and lower-capital letters".format(string1))

    if (string2 != "") and description_col_request == 100000:
        messagebox.showinfo('Info',  r"Requested column '{}' was not found. Check spaces and lower-capital letters".format(string2))

    if (string3 != "") and partno_col_request == 100000:
        messagebox.showinfo('Info',  r"Requested column '{}' was not found. Check spaces and lower-capital letters".format(string3))

    quantity_col_database = 100000

    if column_9_select == 0:
        for a in request_sheet.iter_rows():
            for b in request_sheet.iter_cols():
                col_no = b[0].column -1
                compare = a[col_no].value

                if compare == column_9 and column_9 != "":
                    quantity_col_database = col_no
                    break

    if column_9_select == 1:
        for a in database_sheet.iter_rows():
            for b in database_sheet.iter_cols():
                col_no = b[0].column -1
                compare = a[col_no].value

                if compare == column_9 and column_9 != "":
                    quantity_col_database = col_no
                    break

    description_col_database = 100000
    description_row_database = 100000

    for a in database_sheet.iter_rows():
        row_no = a[0].row
        for b in database_sheet.iter_cols():
            col_no = b[0].column -1
            compare = a[col_no].value

            if string4 == compare and string4 != "" and string4 != " ":
                description_col_database = col_no
                description_row_database = row_no
                break

    if (string4 != "") and description_col_database == 100000:
        messagebox.showinfo('Info',  r"Database column '{}' was not found. Check spaces and lower-capital letters".format(string4))

    QIP_col_database = 100000

    if column_7_select == 0:
        for a in request_sheet.iter_rows():
            for b in request_sheet.iter_cols():
                col_no = b[0].column -1
                compare = a[col_no].value

                if compare == column_7 and column_7 != "":
                    QIP_col_database = col_no
                    break

    if column_7_select == 1:
        for a in database_sheet.iter_rows():
            for b in database_sheet.iter_cols():
                col_no = b[0].column -1
                compare = a[col_no].value

                if compare == column_7 and column_7 != "":
                    QIP_col_database = col_no
                    break

    m = 1
    n = 1
    check = 0
    last_placed = " "
    previous_stored = " "
    Token_Set_Ratio = 0
    Token_Set_Ratio2 = 0
    Token_Set_Ratio3 = 0

    if int(stop_line) == 100000:
        for i in request_sheet.iter_rows(min_row=description_row_request+1):
            if description_col_request < 100000:
                request_component = i[description_col_request].value
            if details_col_request < 100000:
                details_component = i[details_col_request].value
            if partno_col_request < 100000:
                partno_component = i[partno_col_request].value
            row_number = i[0].row
            found = 0
            if description_col_database<100000:
                for j in database_sheet.iter_rows(min_row=description_row_database+1):
                     #if j[0].value == id:
                        #print(j[0].value)
                        database_component = j[description_col_database].value

                        if description_col_request < 100000:
                            if request_component == "0" or request_component == "" or request_component == " ":
                                Token_Set_Ratio = 0
                            else:
                                Token_Set_Ratio = fuzz.token_set_ratio(request_component,database_component)
                        if details_col_request < 100000:
                            if details_component == "0" or details_component == "" or details_component == " ":
                                Token_Set_Ratio2 = 0
                            else:
                                Token_Set_Ratio2 = fuzz.token_set_ratio(details_component,database_component)
                        if partno_col_request < 100000:
                            if partno_component == "0" or partno_component == "" or partno_component == "":
                                Token_Set_Ratio3 = 0
                            else:
                                Token_Set_Ratio3 = fuzz.token_set_ratio(partno_component,database_component)

                        if Token_Set_Ratio > int(match_score) or Token_Set_Ratio2 > int(match_score) or Token_Set_Ratio3 > int(match_score):
                            check = 1
                            if database_component != previous_stored:
                                found = 1
                                m = m+1
                                results_sheet.cell(row=m, column=1).value = row_number

                                if QIP_col_database == 100000 and details_col_request == 100000:
                                    results_sheet.cell(row=m, column=6).value = j[description_col_database].value
                                    results_sheet.cell(row=m, column=10).value = max(Token_Set_Ratio, Token_Set_Ratio2, Token_Set_Ratio3)
                                    if quantity_col_database < 100000:
                                        if column_9_select == 0:
                                            results_sheet.cell(row=m, column=9).value = i[quantity_col_database].value
                                        if column_9_select == 1:
                                            results_sheet.cell(row=m, column=9).value = j[quantity_col_database].value
                                    if case_col_request < 100000:
                                        if column_5_select == 0:
                                            results_sheet.cell(row=m, column=5).value = i[case_col_request].value
                                        if column_5_select == 1:
                                            results_sheet.cell(row=m, column=5).value = j[case_col_request].value
                                    if partno_col_request < 100000:
                                        results_sheet.cell(row=m, column=4).value = i[partno_col_request].value
                                    if description_col_request < 100000:
                                        results_sheet.cell(row=m, column=3).value = i[description_col_request].value
                                    if quantity_col_request < 100000:
                                        if column_8_select == 0:
                                            results_sheet.cell(row=m, column=8).value = i[quantity_col_request].value
                                        if column_8_select == 1:
                                            results_sheet.cell(row=m, column=8).value = j[quantity_col_request].value

                                if QIP_col_database < 100000 and details_col_request == 100000:
                                    results_sheet.cell(row=m, column=6).value = j[description_col_database].value
                                    if column_7_select == 0:
                                        results_sheet.cell(row=m, column=7).value = i[QIP_col_database].value
                                    if column_7_select == 1:
                                        results_sheet.cell(row=m, column=7).value = j[QIP_col_database].value
                                    results_sheet.cell(row=m, column=10).value = max(Token_Set_Ratio, Token_Set_Ratio2, Token_Set_Ratio3)
                                    if quantity_col_database < 100000:
                                        if column_9_select == 0:
                                            results_sheet.cell(row=m, column=9).value = i[quantity_col_database].value
                                        if column_9_select == 1:
                                            results_sheet.cell(row=m, column=9).value = j[quantity_col_database].value
                                    if case_col_request < 100000:
                                        if column_5_select == 0:
                                            results_sheet.cell(row=m, column=5).value = i[case_col_request].value
                                        if column_5_select == 1:
                                            results_sheet.cell(row=m, column=5).value = j[case_col_request].value
                                    if partno_col_request < 100000:
                                        results_sheet.cell(row=m, column=4).value = i[partno_col_request].value
                                    if description_col_request < 100000:
                                        results_sheet.cell(row=m, column=3).value = i[description_col_request].value
                                    if quantity_col_request < 100000:
                                        if column_8_select == 0:
                                            results_sheet.cell(row=m, column=8).value = i[quantity_col_request].value
                                        if column_8_select == 1:
                                            results_sheet.cell(row=m, column=8).value = j[quantity_col_request].value


                                if QIP_col_database == 100000 and details_col_request < 100000:
                                    results_sheet.cell(row=m, column=2).value = i[details_col_request].value
                                    results_sheet.cell(row=m, column=6).value = j[description_col_database].value
                                    results_sheet.cell(row=m, column=10).value = max(Token_Set_Ratio, Token_Set_Ratio2, Token_Set_Ratio3)
                                    if quantity_col_database < 100000:
                                        if column_9_select == 0:
                                            results_sheet.cell(row=m, column=9).value = i[quantity_col_database].value
                                        if column_9_select == 1:
                                            results_sheet.cell(row=m, column=9).value = j[quantity_col_database].value
                                    if case_col_request < 100000:
                                        if column_5_select == 0:
                                            results_sheet.cell(row=m, column=5).value = i[case_col_request].value
                                        if column_5_select == 1:
                                            results_sheet.cell(row=m, column=5).value = j[case_col_request].value
                                    if partno_col_request < 100000:
                                        results_sheet.cell(row=m, column=4).value = i[partno_col_request].value
                                    if description_col_request < 100000:
                                        results_sheet.cell(row=m, column=3).value = i[description_col_request].value
                                    if quantity_col_request < 100000:
                                        if column_8_select == 0:
                                            results_sheet.cell(row=m, column=8).value = i[quantity_col_request].value
                                        if column_8_select == 1:
                                            results_sheet.cell(row=m, column=8).value = j[quantity_col_request].value

                                if QIP_col_database < 100000 and details_col_request < 100000:
                                    results_sheet.cell(row=m, column=2).value = i[details_col_request].value
                                    results_sheet.cell(row=m, column=6).value = j[description_col_database].value
                                    if column_7_select == 0:
                                        results_sheet.cell(row=m, column=7).value = i[QIP_col_database].value
                                    if column_7_select == 1:
                                        results_sheet.cell(row=m, column=7).value = j[QIP_col_database].value
                                    results_sheet.cell(row=m, column=10).value = max(Token_Set_Ratio, Token_Set_Ratio2, Token_Set_Ratio3)
                                    if quantity_col_database < 100000:
                                        if column_9_select == 0:
                                            results_sheet.cell(row=m, column=9).value = i[quantity_col_database].value
                                        if column_9_select == 1:
                                            results_sheet.cell(row=m, column=9).value = j[quantity_col_database].value
                                    if case_col_request < 100000:
                                        if column_5_select == 0:
                                            results_sheet.cell(row=m, column=5).value = i[case_col_request].value
                                        if column_5_select == 1:
                                            results_sheet.cell(row=m, column=5).value = j[case_col_request].value
                                    if partno_col_request < 100000:
                                        results_sheet.cell(row=m, column=4).value = i[partno_col_request].value
                                    if description_col_request < 100000:
                                        results_sheet.cell(row=m, column=3).value = i[description_col_request].value
                                    if quantity_col_request < 100000:
                                        if column_8_select == 0:
                                            results_sheet.cell(row=m, column=8).value = i[quantity_col_request].value
                                        if column_8_select == 1:
                                            results_sheet.cell(row=m, column=8).value = j[quantity_col_request].value

                            previous_stored = database_component
                if found==0:
                    #label_8.grid(row=7, column=3)
                    if details_col_request<100000:
                        if last_placed != i[details_col_request].value and i[details_col_request].value != None and i[details_col_request].value != "" and i[details_col_request].value != "Undefined":
                            #print(i[description_col_request].value)
                            n = n+1
                            if details_col_request < 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                not_found_sheet.cell(row=n, column=2).value = i[details_col_request].value
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            if details_col_request == 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            last_placed = i[details_col_request].value

                    if description_col_request<100000 and details_col_request == 100000 and partno_col_request == 100000:
                        if last_placed != i[description_col_request].value and i[description_col_request].value != None and i[description_col_request].value != "" and i[description_col_request].value != "Undefined":
                            #print(i[description_col_request].value)
                            n = n+1
                            if details_col_request < 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                not_found_sheet.cell(row=n, column=2).value = i[details_col_request].value
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            if details_col_request == 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            last_placed = i[description_col_request].value

                    if partno_col_request < 100000 and details_col_request == 100000 and partno_col_request == 100000:
                        if last_placed != i[partno_col_request].value and i[partno_col_request].value != None and i[partno_col_request].value != "" and i[partno_col_request].value != "Undefined":
                            #print(i[description_col_request].value)
                            n = n+1
                            if details_col_request < 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                not_found_sheet.cell(row=n, column=2).value = i[details_col_request].value
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            if details_col_request == 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            last_placed = i[partno_col_request].value

                    if partno_col_request < 100000 and details_col_request == 100000 and partno_col_request < 100000:
                        if last_placed != i[partno_col_request].value and i[partno_col_request].value != None and i[partno_col_request].value != "" and i[partno_col_request].value != "Undefined":
                            #print(i[description_col_request].value)
                            n = n+1
                            if details_col_request < 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                not_found_sheet.cell(row=n, column=2).value = i[details_col_request].value
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            if details_col_request == 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            last_placed = i[partno_col_request].value

    if int(stop_line) < 100000:
        for i in request_sheet.iter_rows(min_row=description_row_request+1, max_row=int(stop_line)):
            if description_col_request < 100000:
                request_component = i[description_col_request].value
            if details_col_request < 100000:
                details_component = i[details_col_request].value
            if partno_col_request < 100000:
                partno_component = i[partno_col_request].value
            row_number = i[0].row
            found = 0
            if description_col_database<100000:
                for j in database_sheet.iter_rows(min_row=description_row_database+1):
                     #if j[0].value == id:
                        #print(j[0].value)
                        database_component = j[description_col_database].value

                        if description_col_request < 100000:
                            if request_component == "0" or request_component == "" or request_component == " ":
                                Token_Set_Ratio = 0
                            else:
                                Token_Set_Ratio = fuzz.token_set_ratio(request_component,database_component)
                        if details_col_request < 100000:
                            if details_component == "0" or details_component == "" or details_component == " ":
                                Token_Set_Ratio2 = 0
                            else:
                                Token_Set_Ratio2 = fuzz.token_set_ratio(details_component,database_component)
                        if partno_col_request < 100000:
                            if partno_component == "0" or partno_component == "" or partno_component == "":
                                Token_Set_Ratio3 = 0
                            else:
                                Token_Set_Ratio3 = fuzz.token_set_ratio(partno_component,database_component)

                        if Token_Set_Ratio > int(match_score) or Token_Set_Ratio2 > int(match_score) or Token_Set_Ratio3 > int(match_score):
                            check = 1
                            if database_component != previous_stored:
                                found = 1
                                m = m+1
                                results_sheet.cell(row=m, column=1).value = row_number

                                if QIP_col_database == 100000 and details_col_request == 100000:
                                    results_sheet.cell(row=m, column=6).value = j[description_col_database].value
                                    results_sheet.cell(row=m, column=10).value = max(Token_Set_Ratio, Token_Set_Ratio2, Token_Set_Ratio3)
                                    if quantity_col_database < 100000:
                                        if column_9_select == 0:
                                            results_sheet.cell(row=m, column=9).value = i[quantity_col_database].value
                                        if column_9_select == 1:
                                            results_sheet.cell(row=m, column=9).value = j[quantity_col_database].value
                                    if case_col_request < 100000:
                                        if column_5_select == 0:
                                            results_sheet.cell(row=m, column=5).value = i[case_col_request].value
                                        if column_5_select == 1:
                                            results_sheet.cell(row=m, column=5).value = j[case_col_request].value
                                    if partno_col_request < 100000:
                                        results_sheet.cell(row=m, column=4).value = i[partno_col_request].value
                                    if description_col_request < 100000:
                                        results_sheet.cell(row=m, column=3).value = i[description_col_request].value
                                    if quantity_col_request < 100000:
                                        if column_8_select == 0:
                                            results_sheet.cell(row=m, column=8).value = i[quantity_col_request].value
                                        if column_8_select == 1:
                                            results_sheet.cell(row=m, column=8).value = j[quantity_col_request].value

                                if QIP_col_database < 100000 and details_col_request == 100000:
                                    results_sheet.cell(row=m, column=6).value = j[description_col_database].value
                                    if column_7_select == 0:
                                        results_sheet.cell(row=m, column=7).value = i[QIP_col_database].value
                                    if column_7_select == 1:
                                        results_sheet.cell(row=m, column=7).value = j[QIP_col_database].value
                                    results_sheet.cell(row=m, column=10).value = max(Token_Set_Ratio, Token_Set_Ratio2, Token_Set_Ratio3)
                                    if quantity_col_database < 100000:
                                        if column_9_select == 0:
                                            results_sheet.cell(row=m, column=9).value = i[quantity_col_database].value
                                        if column_9_select == 1:
                                            results_sheet.cell(row=m, column=9).value = j[quantity_col_database].value
                                    if case_col_request < 100000:
                                        if column_5_select == 0:
                                            results_sheet.cell(row=m, column=5).value = i[case_col_request].value
                                        if column_5_select == 1:
                                            results_sheet.cell(row=m, column=5).value = j[case_col_request].value
                                    if partno_col_request < 100000:
                                        results_sheet.cell(row=m, column=4).value = i[partno_col_request].value
                                    if description_col_request < 100000:
                                        results_sheet.cell(row=m, column=3).value = i[description_col_request].value
                                    if quantity_col_request < 100000:
                                        if column_8_select == 0:
                                            results_sheet.cell(row=m, column=8).value = i[quantity_col_request].value
                                        if column_8_select == 1:
                                            results_sheet.cell(row=m, column=8).value = j[quantity_col_request].value


                                if QIP_col_database == 100000 and details_col_request < 100000:
                                    results_sheet.cell(row=m, column=2).value = i[details_col_request].value
                                    results_sheet.cell(row=m, column=6).value = j[description_col_database].value
                                    results_sheet.cell(row=m, column=10).value = max(Token_Set_Ratio, Token_Set_Ratio2, Token_Set_Ratio3)
                                    if quantity_col_database < 100000:
                                        if column_9_select == 0:
                                            results_sheet.cell(row=m, column=9).value = i[quantity_col_database].value
                                        if column_9_select == 1:
                                            results_sheet.cell(row=m, column=9).value = j[quantity_col_database].value
                                    if case_col_request < 100000:
                                        if column_5_select == 0:
                                            results_sheet.cell(row=m, column=5).value = i[case_col_request].value
                                        if column_5_select == 1:
                                            results_sheet.cell(row=m, column=5).value = j[case_col_request].value
                                    if partno_col_request < 100000:
                                        results_sheet.cell(row=m, column=4).value = i[partno_col_request].value
                                    if description_col_request < 100000:
                                        results_sheet.cell(row=m, column=3).value = i[description_col_request].value
                                    if quantity_col_request < 100000:
                                        if column_8_select == 0:
                                            results_sheet.cell(row=m, column=8).value = i[quantity_col_request].value
                                        if column_8_select == 1:
                                            results_sheet.cell(row=m, column=8).value = j[quantity_col_request].value

                                if QIP_col_database < 100000 and details_col_request < 100000:
                                    results_sheet.cell(row=m, column=2).value = i[details_col_request].value
                                    results_sheet.cell(row=m, column=6).value = j[description_col_database].value
                                    if column_7_select == 0:
                                        results_sheet.cell(row=m, column=7).value = i[QIP_col_database].value
                                    if column_7_select == 1:
                                        results_sheet.cell(row=m, column=7).value = j[QIP_col_database].value
                                    results_sheet.cell(row=m, column=10).value = max(Token_Set_Ratio, Token_Set_Ratio2, Token_Set_Ratio3)
                                    if quantity_col_database < 100000:
                                        if column_9_select == 0:
                                            results_sheet.cell(row=m, column=9).value = i[quantity_col_database].value
                                        if column_9_select == 1:
                                            results_sheet.cell(row=m, column=9).value = j[quantity_col_database].value
                                    if case_col_request < 100000:
                                        if column_5_select == 0:
                                            results_sheet.cell(row=m, column=5).value = i[case_col_request].value
                                        if column_5_select == 1:
                                            results_sheet.cell(row=m, column=5).value = j[case_col_request].value
                                    if partno_col_request < 100000:
                                        results_sheet.cell(row=m, column=4).value = i[partno_col_request].value
                                    if description_col_request < 100000:
                                        results_sheet.cell(row=m, column=3).value = i[description_col_request].value
                                    if quantity_col_request < 100000:
                                        if column_8_select == 0:
                                            results_sheet.cell(row=m, column=8).value = i[quantity_col_request].value
                                        if column_8_select == 1:
                                            results_sheet.cell(row=m, column=8).value = j[quantity_col_request].value

                            previous_stored = database_component
                if found==0:
                    #label_8.grid(row=7, column=3)
                    if details_col_request<100000:
                        if last_placed != i[details_col_request].value and i[details_col_request].value != None and i[details_col_request].value != "" and i[details_col_request].value != "Undefined":
                            #print(i[description_col_request].value)
                            n = n+1
                            if details_col_request < 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                not_found_sheet.cell(row=n, column=2).value = i[details_col_request].value
                                if quantity_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[quantity_col_request].value
                                if case_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[case_col_request].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            if details_col_request == 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                if quantity_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[quantity_col_request].value
                                if case_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[case_col_request].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            last_placed = i[details_col_request].value

                    if description_col_request<100000 and details_col_request == 100000 and partno_col_request == 100000:
                        if last_placed != i[description_col_request].value and i[description_col_request].value != None and i[description_col_request].value != "" and i[description_col_request].value != "Undefined":
                            #print(i[description_col_request].value)
                            n = n+1
                            if details_col_request < 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                not_found_sheet.cell(row=n, column=2).value = i[details_col_request].value
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            if details_col_request == 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            last_placed = i[description_col_request].value

                    if partno_col_request < 100000 and details_col_request == 100000 and partno_col_request == 100000:
                        if last_placed != i[partno_col_request].value and i[partno_col_request].value != None and i[partno_col_request].value != "" and i[partno_col_request].value != "Undefined":
                            #print(i[description_col_request].value)
                            n = n+1
                            if details_col_request < 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                not_found_sheet.cell(row=n, column=2).value = i[details_col_request].value
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            if details_col_request == 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            last_placed = i[partno_col_request].value

                    if partno_col_request < 100000 and details_col_request == 100000 and partno_col_request < 100000:
                        if last_placed != i[partno_col_request].value and i[partno_col_request].value != None and i[partno_col_request].value != "" and i[partno_col_request].value != "Undefined":
                            #print(i[description_col_request].value)
                            n = n+1
                            if details_col_request < 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                not_found_sheet.cell(row=n, column=2).value = i[details_col_request].value
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            if details_col_request == 100000:
                                not_found_sheet.cell(row=n, column=1).value = row_number
                                if not_col_2 < 100000:
                                    not_found_sheet.cell(row=n, column=6).value = i[not_col_2].value
                                if not_col_1 < 100000:
                                    not_found_sheet.cell(row=n, column=5).value = i[not_col_1].value
                                if partno_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=4).value = i[partno_col_request].value
                                if description_col_request < 100000:
                                    not_found_sheet.cell(row=n, column=3).value = i[description_col_request].value
                            last_placed = i[partno_col_request].value

    #check success
    if check==1:
        success = 1
    else:
        success = 0

    #print(success)

    if success == 1:
        label_8.grid_remove()
        label_10.grid_remove()
        label_9.grid(row=11, column=3)
    if success == 0:
        label_8.grid_remove()
        label_9.grid_remove()
        label_10.grid(row=11, column=3)

    #merge cells
    previous_line = 0
    flag = 0
    for i in results_sheet.iter_rows(min_row=1):
        line_number = i[0].value
        row_no = i[0].row
        if line_number == previous_line and flag == 0:
            start_r = row_no-1
            flag = 1
        if line_number != previous_line and flag == 1:
            flag = 0
            end_r = row_no-1
            results_sheet.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
            results_sheet.merge_cells(start_row=start_r, start_column=2, end_row=end_r, end_column=2)
            results_sheet.merge_cells(start_row=start_r, start_column=3, end_row=end_r, end_column=3)
            results_sheet.merge_cells(start_row=start_r, start_column=4, end_row=end_r, end_column=4)
            results_sheet.merge_cells(start_row=start_r, start_column=5, end_row=end_r, end_column=5)
        previous_line = line_number

    #color cells
    if success==1:
        for i in results_sheet.iter_rows(min_row=2):
            score = int(i[9].value)
            row = i[0].row
            if score > 85:
                fill_cell = PatternFill(patternType='solid',
                                fgColor='7FFFD4')
                results_sheet.cell(row=row, column=6).fill = fill_cell
            if score > 70 and score <= 85:
                fill_cell = PatternFill(patternType='solid',
                                fgColor='FFEC8B')
                results_sheet.cell(row=row, column=6).fill = fill_cell
            if score <= 70:
                fill_cell = PatternFill(patternType='solid',
                                fgColor='FFC0CB')
                results_sheet.cell(row=row, column=6).fill = fill_cell

    #center all cells
    for col in results_sheet.columns:
        for cell in col:
            from copy import copy
            alignment_obj = copy(cell.alignment)
            alignment_obj.horizontal = 'center'
            alignment_obj.vertical = 'center'
            cell.alignment = alignment_obj

    for col in not_found_sheet.columns:
        for cell in col:
            from copy import copy
            alignment_obj = copy(cell.alignment)
            alignment_obj.horizontal = 'center'
            alignment_obj.vertical = 'center'
            cell.alignment = alignment_obj
    #save
    if GUI ==1:
        string_in_string = r"{}\Search_results.xlsx".format(destination)
        results_data.save(string_in_string)

        string_in_string2 = r"{}\NOT_found.xlsx".format(destination)
        not_found_data.save(string_in_string2)
        flag_gif=0
        times = times + 1

    if GUI ==0:
        results_data.save(r'C:\Users\nchatzip\cernbox\WINDOWS\Desktop\Search_results.xlsx')
        not_found_data.save(r'C:\Users\nchatzip\cernbox\WINDOWS\Desktop\NOT_found.xlsx')
        flag_gif=0
        times = times + 1

    global runs
    runs = runs + 1
    return runs

#GUI
#"""

root = Tk()

def open_file():
   file = filedialog.askopenfile(mode='r', filetypes=[('Python Files', '*.py')])
   if file:
      content = file.read()
      file.close()
      print("%d characters in this file" % len(content))

def browseFiles1():
    filename1 = filedialog.askopenfilename(initialdir = "/",
                                          title = "Select a File",
                                          filetypes = (("all files",
                                                        "*.*"),
                                                       ("Text files",
                                                        "*.txt*")))
    textBox1.insert(END,filename1)

def browseFiles2():
    filename2 = filedialog.askopenfilename(initialdir = "/",
                                          title = "Select a File",
                                          filetypes = (("all files",
                                                        "*.*"),
                                                       ("Text files",
                                                        "*.txt*")))
    textBox2.insert(END,filename2)

def browseFiles3():
    filename3 = filedialog.askdirectory()

    textBox3.insert(END,filename3)

def onClick():
    tkinter.messagebox.showinfo("Help me",  "1) Make sure that there are no column titles with same names in excel input files (eg. only one column named 'Description') \n\n"
                                            "2) Search app searches on the first excel sheet of the input files\n\n"
                                            "3) You can use less than 3 request string columns (eg. 1 or 2). Leave empty box(es)\n\n"
                                            "4) Download database file from EAM --> Materials --> Setup --> Stores --> Stock\n")

def stopinfo():
    tkinter.messagebox.showinfo("Stop line info",  "Select a specific line in request file, after which the search will stop. For example you can use this in order to ignore not mounted components which are usually placed on the last lines")

def score_explain():
    tkinter.messagebox.showinfo("String match score",  "It is a score value (0-100) that declares how similar 2 compared strings are. If "
                                                       "you select a higher value, less and more accurate results will be produced. Suggested value is 60.")

entry_1 = Entry(root)
entry_2 = Entry(root)

label_1 = customtkinter.CTkLabel(root, text="Request excel file (.xlsx) location")
label_2 = customtkinter.CTkLabel(root, text="Database excel file (.xlsx) location")
label_3 = customtkinter.CTkLabel(root, text="Results file destination path")
label_4 = customtkinter.CTkLabel(root, text="String match score (0-100)")
label_5 = customtkinter.CTkLabel(root, text=" ")
label_6 = customtkinter.CTkLabel(root, text="Request file search column(s)")
label_7 = customtkinter.CTkLabel(root, text="Database file search column")
label_8 = customtkinter.CTkLabel(root, text="Running..",text_color='blue')
label_9 = customtkinter.CTkLabel(root, text="Success!",text_color='blue')
label_10 = customtkinter.CTkLabel(root, text="No results found..",text_color='blue')
label_11 = customtkinter.CTkLabel(root, text="to results file column")
label_12 = customtkinter.CTkLabel(root, text="to results file column")
label_13 = customtkinter.CTkLabel(root, text="to results file column")
label_14 = customtkinter.CTkLabel(root, text="to results file column")
label_15 = customtkinter.CTkLabel(root, text="Print from request file column (*)")
label_16 = customtkinter.CTkLabel(root, text="Print from request file column (*)")
label_17 = customtkinter.CTkLabel(root, text="Print from database file column (*)")
label_18 = customtkinter.CTkLabel(root, text="Print from database file column (*)")
label_20 = customtkinter.CTkLabel(root, text="(*) Optional textboxes", text_color='blue')

label_1.grid(row=0,padx=8)
label_2.grid(row=1,padx=8)
label_3.grid(row=2,padx=8)
label_4.grid(row=3,padx=8)
label_5.grid(column=1, padx=8)
#label_5.grid(column=2)
label_6.grid(row=4, padx=8)
label_7.grid(row=5, padx=8)
label_5.grid(row=12)
label_11.grid(row=6, column=2, padx=8)
label_12.grid(row=7, column=2, padx=8)
label_13.grid(row=8, column=2, padx=8)
label_14.grid(row=9, column=2, padx=8)
label_15.grid(row=6, padx=8)
label_16.grid(row=7, padx=8)
label_17.grid(row=8, padx=8)
label_18.grid(row=9, padx=8)

label_20.grid(row=11, column=3)

#textBox1=Text(root, height=1, width=15)
#textBox2=Text(root, height=1, width=15)
#textBox3=Text(root, height=1, width=15)
#textBox4old=Text(root, height=1, width=3)
#textBox5old=Text(root, height=1, width=12)
#textBox6old=Text(root, height=1, width=12)
#textBox7old=Text(root, height=1, width=12)
#textBox8old=Text(root, height=1, width=12)

textBox1=customtkinter.CTkTextbox(root, height=1, width=120, border_width=2, corner_radius=5)
textBox2=customtkinter.CTkTextbox(root, height=1, width=120, border_width=2, corner_radius=5)
textBox3=customtkinter.CTkTextbox(root, height=1, width=120, border_width=2, corner_radius=5)
textBox4=customtkinter.CTkTextbox(root, height=1, width=38, border_width=2, corner_radius=5)
textBox5=customtkinter.CTkTextbox(root, height=1, width=100, border_width=2, corner_radius=5)
textBox6=customtkinter.CTkTextbox(root, height=1, width=100, border_width=2, corner_radius=5)
textBox7=customtkinter.CTkTextbox(root, height=1, width=100, border_width=2, corner_radius=5)
textBox8=customtkinter.CTkTextbox(root, height=1, width=100, border_width=2, corner_radius=5)

textBox20=customtkinter.CTkTextbox(root, height=1, width=80, border_width=2, corner_radius=5)
textBox21=customtkinter.CTkTextbox(root, height=1, width=80, border_width=2, corner_radius=5)
textBox22=customtkinter.CTkTextbox(root, height=1, width=80, border_width=2, corner_radius=5)
textBox23=customtkinter.CTkTextbox(root, height=1, width=80, border_width=2, corner_radius=5)

textBox1.grid(row=0, column=1, pady=2)
textBox2.grid(row=1, column=1, pady=2)
textBox3.grid(row=2, column=1, pady=2)
textBox4.grid(row=3, column=1, pady=2)
textBox5.grid(row=4, column=1, padx=6, pady=3)
textBox6.grid(row=4, column=2, padx=6, pady=3)
textBox7.grid(row=4, column=3, padx=24, pady=3)
textBox8.grid(row=5, column=1, padx=6, pady=3)
textBox20.grid(row=6, column=1, padx=6, pady=3)
textBox21.grid(row=7, column=1, padx=6, pady=3)
textBox22.grid(row=8, column=1, padx=6, pady=3)
textBox23.grid(row=9, column=1, padx=6, pady=3)

textBox4.insert(INSERT, "60")

attri_button=customtkinter.CTkButton(root, corner_radius=8, fg_color="grey83", hover_color="white", height=35, width=65, text="Run code ", command=lambda:run_code())
attri_button.grid(row=13, column=1, pady=10)

browse1_button=customtkinter.CTkButton(root, corner_radius=8, fg_color="grey83", hover_color="white", height=1, width=9, text="Browse", command=lambda:browseFiles1())
browse1_button.grid(row=0, column=2)

browse2_button=customtkinter.CTkButton(root, corner_radius=8, fg_color="grey83", hover_color="white",height=1, width=9, text="Browse", command=lambda:browseFiles2())
browse2_button.grid(row=1, column=2)

browse3_button=customtkinter.CTkButton(root, corner_radius=8, fg_color="grey83", hover_color="white",height=1, width=9, text="Browse", command=lambda:browseFiles3())
browse3_button.grid(row=2, column=2)

browse4_button=customtkinter.CTkButton(root, corner_radius=8, fg_color="grey83", hover_color="white",height=1, width=11, text="Info", command=lambda:score_explain())
browse4_button.grid(row=3, column=2)

help_button=customtkinter.CTkButton(root, corner_radius=8, fg_color="grey83", hover_color="white",height=1, width=8, text="Help me !", command=lambda:onClick())
help_button.grid(row=13, column=0)

help_button=customtkinter.CTkButton(root, corner_radius=8, fg_color="grey83", hover_color="white",height=1, width=8, text="Info", command=lambda:stopinfo())
help_button.grid(row=11, column=2)

var1 = tk.IntVar()
c1 = customtkinter.CTkCheckBox(root, text='Stop search after line in request file',variable=var1, onvalue=1, offvalue=0,command=lambda:ignore())
c1.grid(row=11, column=0, pady=2, padx=6)

var2 = tk.IntVar()
c2 = customtkinter.CTkCheckBox(root, text='Disable Garfield gif',variable=var2, onvalue=1, offvalue=0,command=lambda:disable_cat())
c2.grid(row=12, column=0, pady=2, padx=6)

var3 = tk.IntVar()
c3 = customtkinter.CTkCheckBox(root, text='Default strings',variable=var3, onvalue=1, offvalue=0,command=lambda:autocomplete())
c3.grid(row=12, column=1, pady=2, padx=6)

textBox9=customtkinter.CTkTextbox(root, height=1, width=45, border_width=2, corner_radius=5)

# Dropdown menu options
options = [
    "5",
    "7",
    "8",
    "9",
]

# Create Dropdown menu
drop1 = customtkinter.CTkComboBox(root, width=58, values=options)
drop1.set(value="5")
drop1.grid(row=6, column=3, padx=8)

drop2 = customtkinter.CTkComboBox(root, width=58, values=options )
drop2.set(value="8")
drop2.grid(row=7, column=3, padx=8)

drop3 = customtkinter.CTkComboBox(root, width=58, values=options )
drop3.set(value="7")
drop3.grid(row=8, column=3, padx=8)

drop4 = customtkinter.CTkComboBox(root, width=58, values=options )
drop4.set(value="9")
drop4.grid(row=9, column=3, padx=8)

def autocomplete():
    if var3.get() == 1:
        textBox5.insert(INSERT, "Description")
        textBox6.insert(INSERT, "Val&Device")
        textBox7.insert(INSERT, "Part Number")
        textBox8.insert(INSERT, "Description")
        textBox20.insert(INSERT, "Case")
        textBox21.insert(INSERT, "Qty")
        textBox22.insert(INSERT, "Part")
        textBox23.insert(INSERT, "Quantity")

    if var3.get() == 0:
        textBox5.delete("1.0","end")
        textBox6.delete("1.0","end")
        textBox7.delete("1.0","end")
        textBox8.delete("1.0","end")
        textBox20.delete("1.0","end")
        textBox21.delete("1.0","end")
        textBox22.delete("1.0","end")
        textBox23.delete("1.0","end")

def ignore():
    if var1.get() == 1:
        textBox9.insert(INSERT, "500")
        textBox9.grid(row=11, column=1)
        global ignore_line
        ignore_line = 1
        return ignore_line

    if var1.get() == 0:
        if ignore_line == 1:
            textBox9.delete("1.0","end")
            ignore_line = 0
            return ignore_line

def disable_cat():
    if var2.get() == 1:
        global disable_gif
        disable_gif = 1
        #print(disable_gif)
        return disable_gif

    if var2.get() == 0:
        if disable_gif == 1:
            disable_gif = 0
            #print(disable_gif)
            return disable_gif


photo = PhotoImage(file="cern.png")

labelcern = Label(root, image = photo)
labelcern.image = photo
labelcern.grid(row=0, column=3, rowspan=4)

root.title('CERN Search App')
photo = PhotoImage(file = "search.png")
root.iconphoto(False, photo)

# Pop up in the middle of the screen
root.eval('tk::PlaceWindow . center')

frameCnthi = 12
frameshi = [PhotoImage(file='hi.gif',format = 'gif -index %i' %(i)) for i in range(frameCnthi)]

def update2(ind=None):

    if disable_gif == 0:
        framehi = frameshi[ind]
        ind += 1
        if ind >= frameCnthi:
            ind = 0
        labelhi.configure(image=framehi)
        root.after(300, update2, ind)

    if disable_gif == 1:
        photo2 = PhotoImage(file="disabled.png")
        labelcern2 = Label(root, image = photo2)
        labelcern2.image = photo2
        labelcern2.grid(row=12, column=2, rowspan=2, columnspan=3)


labelhi = Label(root)
labelhi.grid(row=12, column=2, rowspan=2, columnspan=3)

if begin_gif==0:
    root.after(0, update2, 0)

root.mainloop()
#"""

#if __name__ == '_main_':
 #   main()
